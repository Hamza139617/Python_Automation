import openpyxl as xl
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
import sys
from pathlib import Path




"""
Scenario : 

A mid-sized retail brand runs both online and physical store sales across Pakistan. Every month, the operations manager gets two separate Excel files from the e-commerce team and the store billing team. Each file is full of raw transaction data, but the manager still has to manually build a summary report showing:

total sales by product category
monthly sales trends
top-performing sales reps
revenue and profit by region

Because the source files change every week, the same reporting work has to be repeated from scratch each month. They want a Python automation solution that can take the raw data and generate a clean summary report automatically.

Your task

Use Python with openpyxl to create an automated reporting workflow that:

reads both raw data sheets,
combines and cleans the data,
builds pivot-style summaries,
formats the output into a professional report sheet,
highlights top categories and top performers,
keeps the process reusable when new data is added.

The workbook includes 2 filled data sheets with lots of realistic sales records, so it is ready to use as the source file for that automation exercise.

"""

### real implementation starts from here 

def safe_int(vale):

    try:
        return int(float(str(vale))) if vale is not None else 0
    except (ValueError, TypeError):
        return 0
    

def create_report_sheet(work_book, category_total_sales, monthly_sales_trends, top_sales_reps, region_revenue, region_profit):

    work_book = Workbook()
    data_ws = work_book.active
    data_ws.sheet_state = "hidden"
    
    data_ws["A1"] = "Category"
    data_ws["B1"] = "Sales"

    for i, item in enumerate(category_total_sales):
        data_ws.cell(row=i + 2, column=1, value=item[0])
        data_ws.cell(row=i + 2, column=2, value=safe_int(item[1]))
    cat_end = len(category_total_sales) + 1

    monthly_sorted = sorted(monthly_sales_trends, key= lambda x: int(x[0]))
    data_ws["D1"] = "Month"
    data_ws["E1"] = "Revenue"

    for i, item in enumerate(monthly_sorted):
        data_ws.cell(row=i +2, column=4, value=item[0])
        data_ws.cell(row=i + 2, column=5, value=safe_int(item[1]))
    month_end = len(monthly_sorted) + 1

    reps_sorted = sorted(top_sales_reps, key=lambda x: int(x[1]), reverse=True)
    data_ws["G1"] = "Sales Rep"
    data_ws["H1"] = "Revenue"

    for i, item in enumerate(reps_sorted):
        data_ws.cell(row=i + 2, column=7, value=item[0])
        data_ws.cell(row=i + 2, column=8, value=safe_int(item[1]))
    reps_end = len(reps_sorted) + 1

    data_ws["J1"] = "Region"
    data_ws["K1"] = "Revenue"
    for i, item in enumerate(region_revenue):
        data_ws.cell(row=i + 2, column=10, value=item[0])
        data_ws.cell(row=i + 2, column= 11, value=safe_int(item[1]))
    reg_rev_end = len(region_revenue) + 1

    data_ws["M1"] = "Region"
    data_ws["N1"] = "Profit"

    for i, item in enumerate(region_profit):
        data_ws.cell(row= i + 2, column=13, value=item[0])
        data_ws.cell(row=i + 2, column=14, value=safe_int(item[1]))
    reg_profit_end = len(region_profit) + 1

    report_ws = work_book.create_sheet("Sales Report", 0)

    def section_label(row_num, text):

        report_ws.merge_cells(f"A{row_num}:T{row_num}")
        c = report_ws.cell(row=row_num, column=1, value=text)
        c.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        report_ws.row_dimensions[row_num].height = 22

    def make_bar(title, y_label, x_label, data_ref, cats_ref, width=15, height=10, horizontal=False):

        c = BarChart()
        c.type = "bar" if horizontal else "col"
        c.title = title
        c.y_axis.title = y_label
        c.x_axis.title = x_label
        c.style = 10
        c.width = width
        c.height = height
        c.add_data(data_ref, titles_from_data=True)
        c.set_categories(cats_ref)
        return c
    
    def make_pie(title, data_ref, cats_ref, width = 15, height = 10):

        c = PieChart()
        c.title = title
        c.style = 10
        c.width = width
        c.height = height
        c.add_data(data_ref, titles_from_data=True)
        c.set_categories(cats_ref)

        return c
    
    report_ws.merge_cells("A1:T1")
    t = report_ws["A1"]
    t.value = "Sales Performance Dashboard"
    t.font = Font(name="Calibri", size=22, bold=True, color = "FFFFFF")
    t.fill = PatternFill("solid", fgColor="1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    report_ws.row_dimensions[1].height = 45

    section_label(3, " Sales by Product Category")

    report_ws.add_chart(
        make_bar(
            "Total Sales by Category", "Revenue , (PKR)", "Category",
            Reference(data_ws, min_col=2, min_row=1, max_row=cat_end),
            Reference(data_ws, min_col=1, min_row=2, max_row=cat_end)
        ),
        "A4"
    )

    
    report_ws.add_chart(
        make_pie(
            "Category Sales Share",
            Reference(data_ws, min_col=2, min_row=1, max_row=cat_end),
            Reference(data_ws, min_col=1, min_row=2, max_row=cat_end)
        ),
        "K4"
    )

    section_label(26, "  Monthly Sales Trends")
 
    report_ws.add_chart(
        make_bar(
            "Monthly Revenue", "Revenue (PKR)", "Month",
            Reference(data_ws, min_col=5, min_row=1, max_row=month_end),
            Reference(data_ws, min_col=4, min_row=2, max_row=month_end),
            width=30, height=10        # wide – spans full row
        ),
        "A27"
    )
 
 
    section_label(48, "  Top Sales Representatives  &  Revenue by Region")
 
    report_ws.add_chart(
        make_bar(
            "Top Sales Representatives", "Sales Rep", "Revenue (PKR)",
            Reference(data_ws, min_col=8, min_row=1, max_row=reps_end),
            Reference(data_ws, min_col=7, min_row=2, max_row=reps_end),
            horizontal=True             # flipped to horizontal bars
        ),
        "A49"
    )
 
    report_ws.add_chart(
        make_bar(
            "Revenue by Region", "Revenue (PKR)", "Region",
            Reference(data_ws, min_col=11, min_row=1, max_row=reg_rev_end),
            Reference(data_ws, min_col=10, min_row=2, max_row=reg_rev_end)
        ),
        "K49"
    )
 
 
    section_label(70, "  Regional Revenue & Profit Distribution")
 
    report_ws.add_chart(
        make_pie(
            "Regional Revenue Share",
            Reference(data_ws, min_col=11, min_row=1, max_row=reg_rev_end),
            Reference(data_ws, min_col=10, min_row=2, max_row=reg_rev_end)
        ),
        "A71"
    )
 
    report_ws.add_chart(
        make_bar(
            "Profit by Region", "Profit (PKR)", "Region",
            Reference(data_ws, min_col=14, min_row=1, max_row=reg_profit_end),
            Reference(data_ws, min_col=13, min_row=2, max_row=reg_profit_end)
        ),
        "K71"
    )

    work_book.save("professional_report.xlsx")
 
 
 




def    pivot_style_summary(category_total_sales, monthly_sales_trends, top_sales_reps, region_revenue, region_profit):
    
    work_book = Workbook()
    work_sheet = work_book.active
    work_sheet.title = "Pivot style summary"

    row = 1
    pos = 1

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet["A1"] = "category total sales"
    pos +=1

    


    for row in range(0, len(category_total_sales)):
        if len(category_total_sales[row]) == 0 :
            continue
        work_sheet["A" + str(pos)] = category_total_sales[row][0]
        work_sheet["C" + str(pos)] = category_total_sales[row][1]
        pos += 1
    
    pos+= 2

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet[f"A{pos}"] = "monthly sales "
    pos +=1

    for row in range(0, len(monthly_sales_trends)):
        if len(monthly_sales_trends[row]) == 0 :
            continue
        work_sheet["A" + str(pos)] = monthly_sales_trends[row][0]
        work_sheet["C" + str(pos)] = monthly_sales_trends[row][1]
        pos += 1

    pos+= 2

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet[f"A{pos}"] = "monthly sales trends"
    pos +=1

    
    monthly_sales_trends.sort(key= lambda x: int(x[0]))

    for row in range(0, len(monthly_sales_trends)):
        if(len(monthly_sales_trends[row]) == 0):
            continue
        if(row == 0 ):
            work_sheet["A" + str(pos)] = monthly_sales_trends[row][0]
            work_sheet["C" + str(pos)] = 0
        else:
            work_sheet["A" + str(pos)] = monthly_sales_trends[row][0]
            work_sheet["C" + str(pos)] = monthly_sales_trends[row][1] - monthly_sales_trends[row-1][1]
        
        pos += 1

    pos+= 2

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet[f"A{pos}"] = "Top sales representatives"
    pos +=1

    top_sales_reps.sort(key = lambda x: int(x[1]))

    sale_index = len(top_sales_reps) - 1

    while(sale_index >= 0):
        if(len(top_sales_reps) == 0):
            continue

        
        work_sheet["A" + str(pos)] = top_sales_reps[sale_index][0]
        work_sheet["C" + str(pos)] = top_sales_reps[sale_index][1]
        sale_index -= 1
        pos += 1
    
    pos+= 2

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet[f"A{pos}"] = "Region Revenue"
    pos +=1

    for row in range(0, len(region_revenue)):
        if len(region_revenue[row]) == 0 :
            continue
        work_sheet["A" + str(pos)] = region_revenue[row][0]
        work_sheet["C" + str(pos)] = region_revenue[row][1]
        pos += 1

    pos+= 2

    work_sheet.merge_cells(f"A{pos}:B{pos}")
    work_sheet[f"A{pos}"] = "Region Profit"
    pos +=1

    for row in range(0, len(region_profit)):
        if len(region_revenue[row]) == 0:
            continue
        work_sheet["A" + str(pos)] = region_profit[row][0]
        work_sheet["C" + str(pos)] = region_profit[row][1]
        pos += 1

    
    create_report_sheet(
        work_book,
        category_total_sales,
        monthly_sales_trends,
        top_sales_reps,
        region_revenue,
        region_profit
    )

    
    work_book.save("pivot_Style_Summary.xlsx")


def     populate_Lists(category_total_sales, monthly_sales_trends, top_sales_reps, region_revenue, region_profit, work_sheet_online, work_sheet_store ):

    row = 2
    

    category_check = False
    trends_check = False
    reps_check  = False
    region_revenue_check = False
    region_profit_check = False
    

    for row in range(2, work_sheet_online.max_row + 1):

        revenue = int((work_sheet_online["K" + str(row)]).value)
        
        for data in category_total_sales:
            if (work_sheet_online["F" + str(row)]).value in data:
                data[1] += revenue
                category_check = True
                break
        
        if(not(category_check)):
            category_total_sales.append([(work_sheet_online["F" + str(row)]).value, revenue])
        
        category_check = False


        # till here we wrote the operation for populating the category total sales list 

        month = ((str((work_sheet_online["B" + str(row)]).value)).split("-"))[1]

        for data in monthly_sales_trends:

            

            if month in data:

                data[1] += revenue
                trends_check = True
                break
        
        if(not(trends_check)):
            monthly_sales_trends.append([month, revenue])
        
        trends_check = False

        # till here we perform the operation of populating the category monthly sales trends list
            
        for data in top_sales_reps:

            if (work_sheet_online["H" + str(row)]).value in data:
                data[1] += revenue
                reps_check = True
                break

        if(not(reps_check)):
            top_sales_reps.append([(work_sheet_online["H" + str(row)]).value, revenue])
        reps_check = False


        for data in region_revenue:

            if (work_sheet_online["C" + str(row)]).value in data:
                data[1] += revenue
                region_revenue_check = True
                break
        if(not(region_revenue_check)):
            region_revenue.append([(work_sheet_online["C" + str(row)]).value, revenue])
        region_revenue_check = False

        
        for data in region_profit:

            if (work_sheet_online["C" + str(row)]).value in data:
                data[1] += int((work_sheet_online["M" + str(row)]).value)
                region_profit_check = True
                break
            
        if(not(region_profit_check)):
            region_profit.append([(work_sheet_online["C" + str(row)]).value, (work_sheet_online["M" + str(row)]).value ])
            
        region_profit_check = False
    


    row = 2
    

    category_check = False
    trends_check = False
    reps_check  = False
    region_revenue_check = False
    region_profit_check = False


    for row in range(2, work_sheet_store.max_row + 1):

        revenue = int((work_sheet_store["K" + str(row)]).value)
        
        for data in category_total_sales:
            if (work_sheet_store["F" + str(row)]).value in data:
                data[1] += revenue
                category_check = True
                break
        
        if(not(category_check)):
            category_total_sales.append([(work_sheet_store["F" + str(row)]).value, revenue])
        
        category_check = False


        # till here we wrote the operation for populating the category total sales list 

        month = ((str((work_sheet_store["B" + str(row)]).value)).split("-"))[1]

        for data in monthly_sales_trends:

            

            if month in data:

                data[1] += revenue
                trends_check = True
                break
        
        if(not(trends_check)):
            monthly_sales_trends.append([month, revenue])
        
        trends_check = False

        # till here we perform the operation of populating the category monthly sales trends list
            
        for data in top_sales_reps:

            if (work_sheet_store["H" + str(row)]).value in data:
                data[1] += revenue
                reps_check = True
                break

        if(not(reps_check)):
            top_sales_reps.append([(work_sheet_store["H" + str(row)]).value, revenue])
        reps_check = False


        for data in region_revenue:

            if (work_sheet_store["C" + str(row)]).value in data:
                data[1] += revenue
                region_revenue_check = True
                break
        if(not(region_revenue_check)):
            region_revenue.append([(work_sheet_store["C" + str(row)]).value, revenue])
        region_revenue_check = False

        
        for data in region_profit:

            if (work_sheet_store["C" + str(row)]).value in data:
                data[1] += int((work_sheet_store["M" + str(row)]).value)
                region_profit_check = True
                break
            
        if(not(region_profit_check)):
            region_profit.append([(work_sheet_store["C" + str(row)]).value, int((work_sheet_store["M" + str(row)]).value) ])
            
        region_profit_check = False

    

    
    

  




            

            

            





def excel_Automation( path ):

    if path == None:
        print("Please enter the path ")
        sys.exit(1)
    
    
    file = Path(path)

    print(file)
    print(path)

    if(file.exists()):
        print(" file exists ")
    else:
        print("file doesn't exis ")
        sys.exit(1)
    
    work_book = xl.load_workbook(path)

    try:
        work_sheet_online = work_book["Online_Orders"]
    except Exception as e:
        print(e)
    
    try:
        work_sheet_store = work_book["Store_Orders"]
    except Exception as e:
        print(e)
    

    ## data list creation 

    category_total_sales = [] 
    monthly_sales_trends = []
    top_sales_reps = []
    region_revenue = []
    region_profit = []
    

    populate_Lists(
        category_total_sales, monthly_sales_trends, top_sales_reps,
        region_revenue, region_profit, work_sheet_online, work_sheet_store
    )

    pivot_style_summary(
        category_total_sales, monthly_sales_trends, top_sales_reps,
        region_revenue, region_profit
    )



if __name__ == "__main__":
    excel_Automation(sys.argv[1])

